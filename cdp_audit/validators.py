"""Validation helpers."""

import ipaddress
import logging
import socket
from pathlib import Path
from typing import List

import openpyxl

from .app_config import config

logger = logging.getLogger(__name__)


def validate_required_files(paths: List[Path]) -> None:
    """Ensure required files exist before execution.

    Raises:
        SystemExit: If any file paths are missing.
    """
    missing = [str(p) for p in paths if not p.exists()]
    if missing:
        logger.error("Required files missing: %s", ", ".join(missing))
        raise SystemExit(1)


def validate_excel_template(template_path: Path) -> None:
    """Validate that the Excel template exists and has required sheet structure.

    Raises:
        SystemExit: If the template is missing or invalid.
    """
    if not template_path.exists():
        logger.error("Excel template not found: %s", template_path)
        raise SystemExit(1)

    try:
        wb = openpyxl.load_workbook(template_path, data_only=False)
        required_sheets = [
            config.EXCEL_SHEET_AUDIT,
            config.EXCEL_SHEET_DNS,
            config.EXCEL_SHEET_AUTH_ERRORS,
            config.EXCEL_SHEET_CONN_ERRORS,
        ]
        missing_sheets = [sheet for sheet in required_sheets if sheet not in wb.sheetnames]

        if missing_sheets:
            logger.error(
                "Excel template is missing required sheets: %s (has: %s)",
                ", ".join(missing_sheets),
                ", ".join(wb.sheetnames),
            )
            wb.close()
            raise SystemExit(1)

        audit_sheet = wb[config.EXCEL_SHEET_AUDIT]
        cell_value = audit_sheet[config.EXCEL_CELL_SITE_NAME].value
        if cell_value is None or cell_value == "":
            logger.warning("Audit sheet may not be properly formatted (B4 seems empty)")

        wb.close()
        logger.debug("Excel template validated successfully: %s", template_path)

    except FileNotFoundError:
        logger.error("Excel template file not readable: %s", template_path)
        raise SystemExit(1)
    except Exception as exc:
        logger.error("Error validating Excel template: %s", exc)
        raise SystemExit(1)


def normalize_seeds(seeds: List[str]) -> List[str]:
    """Validate seed IPs/hostnames and normalize to IP addresses.

    Raises:
        SystemExit: If a seed cannot be parsed or resolved.
    """
    validated_seeds_set = set()
    for seed in seeds:
        try:
            ipaddress.ip_address(seed)
            validated_seeds_set.add(seed)
        except ValueError:
            try:
                resolved = socket.gethostbyname(seed)
                validated_seeds_set.add(resolved)
                logger.debug("Seed hostname '%s' resolved to %s", seed, resolved)
            except Exception as exc:
                logger.error(
                    "Seed '%s' is not a valid IP and could not be resolved: %s. Aborting.",
                    seed,
                    exc,
                )
                raise SystemExit(1)

    if len(validated_seeds_set) < len(seeds):
        removed = len(seeds) - len(validated_seeds_set)
        logger.warning("Removed %d duplicate seed(s). Starting with %d unique devices.", removed, len(validated_seeds_set))

    return list(validated_seeds_set)
